import json
import time
from selenium import webdriver
from collections import OrderedDict
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from random import randint as ri, random
from datetime import datetime
from openpyxl import load_workbook, Workbook


measures_filename = 'routes.xlsx'
config = json.load(open('config.json'))


def parse_time(text):
    parts, mins = text.strip().split(), 0
    if len(parts) == 2:
        assert(parts[1] == 'мин')
        mins += int(parts[0])
    elif len(parts) == 4:
        assert(parts[1] == 'ч' and parts[3] == 'мин')
        mins += int(parts[0]) * 60 + int(parts[2])
    return mins


def prepare_driver(config):
    options = webdriver.ChromeOptions()
    options.binary_location = config['params']['chrome_location']
    driver = webdriver.Chrome(config['params']['chrome_driver'], chrome_options=options)
    driver.implicitly_wait(1 * 60)
    return driver


def route_between(driver, origin, location):
    src_input, dst_input = driver.find_elements_by_class_name('input_medium__control')

    (ActionChains(driver)
        .pause(ri(3, 10))
        .move_to_element(src_input)
        .send_keys(origin)
        .send_keys(Keys.RETURN)
        .perform())
    (ActionChains(driver)
        .move_to_element(dst_input)
        .send_keys(location)
        .send_keys(Keys.RETURN)
        .pause(ri(4, 7))
        .perform())

    routes = driver.find_elements_by_class_name('driving-route-view__route-title-primary')
    times_from_home = [parse_time(r.text) for r in routes]

    (ActionChains(driver)
        .move_to_element(driver.find_element_by_class_name('route-form-view__reverse-icon'))
        .click()
        .pause(ri(4, 7))
        .perform())

    routes = driver.find_elements_by_class_name('driving-route-view__route-title-primary')
    times_to_home = [parse_time(r.text) for r in routes]

    (ActionChains(driver)
        .move_to_element(driver.find_element_by_class_name('route-form-view__reset'))
        .click()
        .perform())

    return (times_from_home, times_to_home)


def get_times(driver, config, sleep_probability=0.15):
    driver.get('https://yandex.ru/maps/213/moscow/?mode=routes&rtext=&rtt=auto')

    data = {'from_home': OrderedDict(), 'to_home': OrderedDict()}
    for loc in config['locations']:
        data['from_home'][loc], data['to_home'][loc] = route_between(driver, config['origin'], loc)
        if random() < sleep_probability:
            time.sleep(ri(10, 30))

    return data


def flatten_times(times):
    if len(times) == 0:
        return [None, None, None]
    elif len(times) == 1:
        return [times[0], None, None]
    elif len(times) == 2:
        return [min(times), None, max(times)]
    elif len(times) == 3:
        return list(sorted(times))
    ts = list(sorted(times))
    return [ts[0], sum(ts[1:-1], 0.0) / len(ts[1:-1]), ts[-1]]


def rows_from_measures(measure_id, time, data):
    rows = []
    for direction in ['from_home', 'to_home']:
        for type_id, type_name in enumerate(['min', 'mid', 'max']):
            row = [time, measure_id, direction, type_name]
            row += [data[direction][loc][type_id] for loc in data[direction].keys()]
            rows.append(row)
    return rows


def last_measure_id(sheet):
    if sheet.max_row <= 1:
        return 0
    v = sheet.cell(row=sheet.max_row, column=2).value
    return v


if __name__ == '__main__':
    driver = prepare_driver(config)

    now = datetime.now()

    data = get_times(driver, config)
    for loc in config['locations']:
        data['from_home'][loc] = flatten_times(data['from_home'][loc])
        data['to_home'][loc]   = flatten_times(data['to_home'][loc])
    
    driver.close()

    try:
        workbook = load_workbook(measures_filename)
    except:
        workbook = Workbook()
        workbook.active.title = 'Measures'
        locs = [l.replace('Москва, ', '') for l in config['locations']]
        workbook.active.append(['Время', 'MeasureID', 'Направление', 'Тип'] + locs)

    sheet = workbook.active
    m_id = last_measure_id(sheet) + 1
    for row in rows_from_measures(m_id, now, data):
        sheet.append(row)

    workbook.save(measures_filename)
